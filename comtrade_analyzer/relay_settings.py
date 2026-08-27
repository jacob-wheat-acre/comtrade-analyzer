#!/usr/bin/env python3
"""
relay_settings.py — The SUBNET relay settings catalog.

SUBNET exports a *flattened device settings table*, not a native SEL settings
file: one row per relay, carrying asset metadata, the template name, the CT
ratio, and phase/ground pickups for up to three setting groups. This module
treats it as a catalog to look devices up in, which is what it is.

Why it matters here
-------------------
EPSS does two things: it disables reclosing, and it makes the relay more
sensitive. The second effect is invisible without settings — a fault that a
downstream fuse cleared today may trip the recloser on a WSO day, dropping the
whole section instead of one lateral. Deciding that needs the normal-day pickup
and the EPSS pickup for the device that recorded the event.

Units
-----
Pickups are exported in **secondary** amps and CTR converts them to primary,
which is what a COMTRADE record measures:

    primary_pickup_A = pickup_secondary_A * CTR

If your export is already in primary, pass ``pickups_are_primary=True`` to
``load_settings`` — ``sanity_check`` will flag the mismatch either way.

Usage
-----
    cat = load_settings("subnet_settings.xlsx")
    s   = cat.lookup("RCL_076-024")
    ev  = s.evaluate(measured_primary_a=612.0)
"""

from __future__ import annotations

import csv
import os
import re
from dataclasses import dataclass, field
from typing import Dict, List, Optional

# Values SUBNET writes when a setting is absent.
_NULL_TOKENS = {"", "not found", "none", "null", "n/a", "na", "-", "--"}

# Header → canonical field. Matching is case-insensitive and ignores spaces and
# underscores, because column headings drift between exports.
_HEADER_MAP = {
    "id": "record_id",
    "name": "relay_name",
    "feeder": "feeder",
    "templatedate": "template_raw",        # holds the template NAME, not a date
    "template": "template_raw",
    "templatetype": "template_type",
    "admstemplateversion": "adms_version",
    "datetime": "timestamp",
    "ctr": "ctr",
    "nominalsg": "nominal_sg",
    "activesg": "active_sg",
}
for _n in (1, 2, 3):
    _HEADER_MAP[f"sg{_n}phase"] = f"sg{_n}_phase"
    _HEADER_MAP[f"sg{_n}ground"] = f"sg{_n}_ground"


def _key(header: str) -> str:
    return re.sub(r"[\s_\-]+", "", (header or "").strip().lower())


def _clean(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return None if text.lower() in _NULL_TOKENS else text


def _number(value) -> Optional[float]:
    text = _clean(value)
    if text is None:
        return None
    text = text.replace(",", "")
    m = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(m.group()) if m else None


def normalize_name(name: str) -> str:
    """Match keys the way the device registry does, so both agree."""
    return re.sub(r"[\s_\-\.]+", "", (name or "").strip().lower())


# ---------------------------------------------------------------------------
# Template name
# ---------------------------------------------------------------------------

@dataclass
class RelayTemplate:
    """
    The template string carries real metadata, e.g.

        SEL-651R-WF-3PhTrip3PhLoc.2
        └─ relay ─┘ └┘ └── modes ──┘ └ version
    """
    raw: str
    relay_type: Optional[str] = None
    application: Optional[str] = None
    trip_mode: Optional[str] = None
    location_mode: Optional[str] = None
    version: Optional[str] = None

    @property
    def is_wildfire(self) -> bool:
        return (self.application or "").upper() in ("WF", "WILDFIRE")


def parse_template(raw: Optional[str]) -> Optional[RelayTemplate]:
    """Decompose a template name, keeping whatever will not parse."""
    text = _clean(raw)
    if text is None:
        return None
    tpl = RelayTemplate(raw=text)

    body, _, version = text.rpartition(".")
    if body and version.isdigit():
        tpl.version = version
    else:
        body = text

    m = re.match(r"^(SEL-?\w+)", body, re.IGNORECASE)
    if m:
        tpl.relay_type = m.group(1)
        body = body[m.end():].lstrip("-")

    parts = body.split("-")
    if parts and parts[0] and not re.search(r"trip|loc", parts[0], re.IGNORECASE):
        tpl.application = parts[0]
        parts = parts[1:]

    # The modes run together — "3PhTrip3PhLoc". Keep the letter run lazy and
    # forbid it crossing a digit, or the Loc match swallows the Trip mode with
    # it and both come back as one token.
    modes = "-".join(parts)
    m = re.search(r"(\d*Ph[A-Za-z]*?Trip)", modes, re.IGNORECASE)
    if m:
        tpl.trip_mode = m.group(1)
    m = re.search(r"(\d*Ph[A-Za-z]*?Loc)", modes, re.IGNORECASE)
    if m:
        tpl.location_mode = m.group(1)
    return tpl


# ---------------------------------------------------------------------------
# Setting groups
# ---------------------------------------------------------------------------

@dataclass
class SettingGroup:
    number: int
    phase_pickup: Optional[float] = None      # as exported
    ground_pickup: Optional[float] = None

    @property
    def is_populated(self) -> bool:
        return self.phase_pickup is not None or self.ground_pickup is not None


@dataclass
class RelaySettings:
    relay_name: str
    feeder: Optional[str] = None
    record_id: Optional[str] = None
    template: Optional[RelayTemplate] = None
    template_type: Optional[str] = None
    adms_version: Optional[str] = None
    timestamp: Optional[str] = None
    ctr: Optional[float] = None
    nominal_sg: Optional[int] = None
    active_sg: Optional[int] = None
    groups: Dict[int, SettingGroup] = field(default_factory=dict)
    pickups_are_primary: bool = False

    # ── group selection ──────────────────────────────────────────────────────

    def group(self, number: Optional[int]) -> Optional[SettingGroup]:
        if number is None:
            return None
        g = self.groups.get(int(number))
        return g if (g and g.is_populated) else None

    def normal_group(self) -> Optional[SettingGroup]:
        """The normal-day group is the nominal one."""
        return self.group(self.nominal_sg)

    def epss_group(self, forced: Optional[int] = None) -> Optional[SettingGroup]:
        """
        The EPSS group.

        Explicit wins. Otherwise take the most sensitive *other* populated
        group — EPSS lowers pickup, so the lowest phase pickup is the
        candidate. This is inferred, not read: `epss_group_source()` says
        which, and callers should surface that rather than let it pass as fact.
        """
        if forced is not None:
            return self.group(forced)
        others = [g for n, g in sorted(self.groups.items())
                  if g.is_populated and n != self.nominal_sg
                  and g.phase_pickup is not None]
        if not others:
            return None
        return min(others, key=lambda g: g.phase_pickup)

    def epss_group_source(self, forced: Optional[int] = None) -> str:
        if forced is not None:
            return f"SG{forced} (configured)"
        g = self.epss_group()
        return f"SG{g.number} (inferred: lowest pickup outside the nominal group)" if g else "unknown"

    # ── pickup arithmetic ────────────────────────────────────────────────────

    def primary_pickup(self, group: Optional[SettingGroup], kind: str = "phase") -> Optional[float]:
        """Pickup in primary amps, which is what a COMTRADE record measures."""
        if group is None:
            return None
        raw = group.phase_pickup if kind == "phase" else group.ground_pickup
        if raw is None:
            return None
        if self.pickups_are_primary:
            return raw
        if not self.ctr:
            return None
        return raw * self.ctr

    def evaluate(self, measured_primary_a: float, kind: str = "phase",
                 epss_group: Optional[int] = None) -> dict:
        """
        Would this current have picked up, on each setting group?

        Returns pickups in primary amps, the multiple of pickup, and — the
        point of the exercise — whether the answer differs between normal and
        EPSS settings.
        """
        normal = self.normal_group()
        epss = self.epss_group(epss_group)

        def one(g):
            pu = self.primary_pickup(g, kind)
            if pu is None or pu <= 0:
                return {"group": g.number if g else None, "pickup_a": None,
                        "multiple": None, "picks_up": None}
            return {"group": g.number, "pickup_a": round(pu, 1),
                    "multiple": round(measured_primary_a / pu, 2),
                    "picks_up": measured_primary_a >= pu}

        n, e = one(normal), one(epss)
        converts = (n["picks_up"] is False and e["picks_up"] is True)
        return {
            "measured_a": round(measured_primary_a, 1),
            "kind": kind,
            "ctr": self.ctr,
            "normal": n,
            "epss": e,
            "epss_group_source": self.epss_group_source(epss_group),
            # The whole reason this module exists.
            "converts_under_epss": converts,
            "resolved": n["picks_up"] is not None and e["picks_up"] is not None,
        }


# ---------------------------------------------------------------------------
# Catalog
# ---------------------------------------------------------------------------

@dataclass
class SettingsCatalog:
    relays: Dict[str, RelaySettings] = field(default_factory=dict)
    source: str = ""
    warnings: List[str] = field(default_factory=list)

    def lookup(self, device_id: str) -> Optional[RelaySettings]:
        return self.relays.get(normalize_name(device_id))

    def __len__(self) -> int:
        return len(self.relays)


def _rows_from_csv(path: str):
    with open(path, newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            yield row


def _rows_from_excel(path: str):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit(
            f"Reading {os.path.basename(path)} needs openpyxl:\n"
            "    python -m pip install openpyxl\n"
            "Or export the sheet as CSV and point at that instead."
        ) from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [str(h) if h is not None else "" for h in next(rows, [])]
    for values in rows:
        if values is None or all(v is None for v in values):
            continue
        yield dict(zip(headers, values))


def load_settings(path: str, pickups_are_primary: bool = False) -> SettingsCatalog:
    """Load a SUBNET settings export (.csv or .xlsx) into a catalog."""
    ext = os.path.splitext(path)[1].lower()
    rows = _rows_from_excel(path) if ext in (".xlsx", ".xlsm") else _rows_from_csv(path)

    cat = SettingsCatalog(source=path)
    seen_headers, unmapped = False, []

    for raw_row in rows:
        row = {}
        for header, value in raw_row.items():
            canon = _HEADER_MAP.get(_key(header))
            if canon:
                row[canon] = value
            elif not seen_headers and _clean(header):
                unmapped.append(str(header))
        if not seen_headers:
            seen_headers = True
            if unmapped:
                cat.warnings.append(
                    f"{len(unmapped)} column(s) not recognised and ignored: "
                    + ", ".join(unmapped[:8]) + ("…" if len(unmapped) > 8 else ""))

        name = _clean(row.get("relay_name"))
        if not name:
            continue

        groups = {}
        for n in (1, 2, 3):
            g = SettingGroup(number=n,
                             phase_pickup=_number(row.get(f"sg{n}_phase")),
                             ground_pickup=_number(row.get(f"sg{n}_ground")))
            if g.is_populated:
                groups[n] = g

        nominal = _number(row.get("nominal_sg"))
        active = _number(row.get("active_sg"))
        cat.relays[normalize_name(name)] = RelaySettings(
            relay_name=name,
            feeder=_clean(row.get("feeder")),
            record_id=_clean(row.get("record_id")),
            template=parse_template(row.get("template_raw")),
            template_type=_clean(row.get("template_type")),
            adms_version=_clean(row.get("adms_version")),
            timestamp=_clean(row.get("timestamp")),
            ctr=_number(row.get("ctr")),
            nominal_sg=int(nominal) if nominal is not None else None,
            active_sg=int(active) if active is not None else None,
            groups=groups,
            pickups_are_primary=pickups_are_primary,
        )
    return cat


# ---------------------------------------------------------------------------
# Catalog health
# ---------------------------------------------------------------------------

def sanity_check(cat: SettingsCatalog) -> List[dict]:
    """
    Findings about the catalog itself, in the diagnostics.py shape.

    A settings file that loads but is missing CTRs, or whose pickups are in the
    wrong units, produces confident and wrong answers about whether EPSS would
    trip. Say so up front.
    """
    out = []

    def add(level, code, message, detail="", fix=""):
        out.append({"level": level, "code": code, "message": message,
                    "detail": detail, "fix": fix, "file": os.path.basename(cat.source)})

    if not cat.relays:
        add("error", "settings_empty", "No relays loaded from the settings export",
            f"Read {cat.source} but found no usable rows.",
            "The loader keys on a 'Name' column. Check the header row is the "
            "first row of the sheet and that the device column is called Name.")
        return out

    for w in cat.warnings:
        add("info", "settings_columns", w, "", "Harmless if those columns are not settings.")

    no_ctr = [s.relay_name for s in cat.relays.values() if not s.ctr]
    if no_ctr:
        add("warn", "settings_no_ctr",
            f"{len(no_ctr)} relay(s) have no CT ratio",
            ", ".join(no_ctr[:6]) + ("…" if len(no_ctr) > 6 else ""),
            "Pickups are in secondary amps; without CTR they cannot be compared "
            "against a COMTRADE record. Those devices will stay unresolved.")

    no_nominal = [s.relay_name for s in cat.relays.values() if s.nominal_sg is None]
    if no_nominal:
        add("warn", "settings_no_nominal",
            f"{len(no_nominal)} relay(s) have no NOMINAL_SG",
            ", ".join(no_nominal[:6]) + ("…" if len(no_nominal) > 6 else ""),
            "Without it the normal-day group is unknown and no comparison is possible.")

    single = [s.relay_name for s in cat.relays.values()
              if len([g for g in s.groups.values() if g.is_populated]) < 2]
    if single:
        add("warn", "settings_one_group",
            f"{len(single)} relay(s) have only one populated setting group",
            ", ".join(single[:6]) + ("…" if len(single) > 6 else ""),
            "EPSS comparison needs a normal group and a more sensitive one. "
            "These devices can be checked against normal settings only.")

    # Units check: secondary pickups are typically single digits.
    pickups = [g.phase_pickup for s in cat.relays.values()
               for g in s.groups.values() if g.phase_pickup is not None]
    if pickups:
        median = sorted(pickups)[len(pickups) // 2]
        if not any(s.pickups_are_primary for s in cat.relays.values()) and median > 100:
            add("warn", "settings_units",
                f"Phase pickups look like primary amps already (median {median:g} A)",
                "They are being multiplied by CTR, which would overstate them "
                "by that factor.",
                "If the export is already primary, pass --settings-primary.")
    return out
